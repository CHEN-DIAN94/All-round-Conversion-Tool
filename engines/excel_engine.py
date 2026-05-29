"""
engines.excel_engine — Excel → 图片转换引擎

基于 openpyxl 读取工作簿，Pillow 渲染为图片。
支持多工作表（纵向拼接）、自动列宽、网格线、表头高亮。
"""

from pathlib import Path
from typing import Callable, Optional

from utils import finalize_file
from engines._common import _prepare_output


# 缓存 Pillow 字体对象（避免每次 Excel→图片转换重复加载）
_font_cache: Optional[tuple] = None  # (font, font_bold)


def _load_font(candidates: list, size: int = 13):
    """
    尝试从候选列表加载字体，返回第一个成功的。
    全部失败则返回 None（由调用方决定 fallback）。
    """
    from PIL import ImageFont
    for name in candidates:
        try:
            return ImageFont.truetype(name, size)
        except (OSError, IOError):
            continue
    return None


def convert_excel_to_image(
    input_path: str,
    output_path: str,
    progress_callback: Optional[Callable[[int], None]] = None,
) -> str:
    """
    Excel 文件转换为图片。

    使用 openpyxl 读取工作簿，Pillow 渲染为图片。
    支持多工作表（纵向拼接）、自动列宽、网格线、表头高亮。
    """
    from openpyxl import load_workbook
    from PIL import Image, ImageDraw, ImageFont

    temp_path = _prepare_output(output_path, min_free_mb=100)

    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        total_sheets = len(sheets)

        # ---------- 渲染参数 ----------
        cell_padding_x = 8
        cell_padding_y = 6
        min_col_width = 60
        max_col_width = 300
        row_height = 28
        header_bg = (0, 120, 215)       # Windows 蓝
        header_fg = (255, 255, 255)
        grid_color = (220, 220, 220)
        bg_color = (255, 255, 255)
        text_color = (31, 31, 31)
        alt_row_bg = (247, 248, 250)
        sheet_label_bg = (240, 240, 240)
        sheet_label_fg = (80, 80, 80)

        # ---------- 加载字体（带缓存） ----------
        global _font_cache
        if _font_cache is not None:
            font, font_bold = _font_cache
        else:
            regular_fonts = ['msyh.ttc', 'simsun.ttc', 'arial.ttf',
                             'DejaVuSans.ttf', 'NotoSansCJKsc-Regular.otf']
            bold_fonts = ['msyhbd.ttc', 'msyh.ttc', 'simsun.ttc', 'arialbd.ttf',
                          'DejaVuSans-Bold.ttf', 'NotoSansCJKsc-Bold.otf']
            font = _load_font(regular_fonts) or ImageFont.load_default()
            font_bold = _load_font(bold_fonts) or font
            _font_cache = (font, font_bold)

        # ---------- 第一遍：计算每个 sheet 的列宽和总尺寸 ----------
        MAX_ROWS = 5000  # 防止超大文件 OOM
        sheet_infos = []
        truncated_sheets = []

        for sheet_name in sheets:
            ws = wb[sheet_name]
            rows_data = []
            max_col = 0
            row_count = 0

            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                if any(v is not None for v in row_data):
                    rows_data.append(row_data)
                    max_col = max(max_col, len(row_data))
                    row_count += 1
                    if row_count >= MAX_ROWS:
                        truncated_sheets.append(sheet_name)
                        break

            if not rows_data:
                rows_data = [['（空工作表）']]
                max_col = 1

            # 使用 openpyxl 的 max_column 作为真实列数（含中间空列）
            actual_max_col = max(max_col, ws.max_column or 1)

            # 计算列宽：基于内容长度
            col_widths = [min_col_width] * actual_max_col
            for row_data in rows_data:
                for ci, val in enumerate(row_data):
                    if val is not None:
                        text = str(val)
                        # 粗略估算：中文字符宽度约为英文的 1.5 倍
                        char_width = sum(1.5 if ord(c) > 127 else 1 for c in text)
                        needed = int(char_width * 8 + cell_padding_x * 2)
                        col_widths[ci] = min(max(col_widths[ci], needed), max_col_width)

            sheet_width = sum(col_widths) + 1
            sheet_height = (row_count + 1) * row_height + 1  # +1 for header

            sheet_infos.append({
                'name': sheet_name,
                'rows_data': rows_data,
                'col_widths': col_widths,
                'max_col': actual_max_col,
                'max_row': row_count,
                'width': sheet_width,
                'height': sheet_height,
            })
    finally:
        wb.close()

    # ---------- 计算总画布尺寸（含上界防御） ----------
    label_height = 36  # 工作表标签高度
    gap = 12           # 工作表之间的间距
    canvas_width = max(si['width'] for si in sheet_infos) + 40
    canvas_height = sum(si['height'] + label_height + gap for si in sheet_infos) + 20

    # FIX-08: 内存保护 — 超限直接拒绝，避免 OOM
    MAX_CANVAS_PIXELS = 4096 * 4096  # 约 48MB (RGB)
    if canvas_width * canvas_height > MAX_CANVAS_PIXELS:
        raise RuntimeError(
            f'Excel 表格过大（渲染尺寸 {canvas_width}×{canvas_height}），'
            f'请减少数据量或拆分工作表后重试。'
        )

    # ---------- 创建画布 ----------
    img = Image.new('RGB', (canvas_width, canvas_height), bg_color)
    try:
        draw = ImageDraw.Draw(img)

        # ---------- 第二遍：渲染每个 sheet ----------
        y_offset = 10
        for si_idx, si in enumerate(sheet_infos):
            # 工作表标签
            draw.rectangle(
                [10, y_offset, canvas_width - 10, y_offset + label_height],
                fill=sheet_label_bg,
            )
            label_text = f'Sheet: {si["name"]}'
            if si['name'] in truncated_sheets:
                label_text += f'  (已截断，最多 {MAX_ROWS} 行)'
            draw.text(
                (20, y_offset + 8),
                label_text,
                fill=sheet_label_fg,
                font=font_bold,
            )
            y_offset += label_height

            col_widths = si['col_widths']
            rows_data = si['rows_data']

            # 渲染表头（第一行）
            x = 10
            header_y = y_offset
            for ci in range(si['max_col']):
                w = col_widths[ci]
                draw.rectangle(
                    [x, header_y, x + w, header_y + row_height],
                    fill=header_bg,
                )
                if ci < len(rows_data[0]) and rows_data[0][ci] is not None:
                    text = str(rows_data[0][ci])
                    # 截断过长文本
                    max_chars = max(3, int((w - cell_padding_x * 2) / 8))
                    if len(text) > max_chars:
                        text = text[:max_chars - 1] + '…'
                    draw.text(
                        (x + cell_padding_x, header_y + cell_padding_y),
                        text,
                        fill=header_fg,
                        font=font_bold,
                    )
                x += w

            # 渲染数据行
            for ri in range(1, len(rows_data)):
                row_data = rows_data[ri]
                ry = header_y + ri * row_height
                x = 10
                row_bg = alt_row_bg if ri % 2 == 0 else bg_color

                for ci in range(si['max_col']):
                    w = col_widths[ci]
                    draw.rectangle(
                        [x, ry, x + w, ry + row_height],
                        fill=row_bg,
                    )
                    if ci < len(row_data) and row_data[ci] is not None:
                        text = str(row_data[ci])
                        max_chars = max(3, int((w - cell_padding_x * 2) / 8))
                        if len(text) > max_chars:
                            text = text[:max_chars - 1] + '…'
                        draw.text(
                            (x + cell_padding_x, ry + cell_padding_y),
                            text,
                            fill=text_color,
                            font=font,
                        )
                    x += w

            # 绘制网格线
            # 垂直线
            x = 10
            for ci in range(si['max_col'] + 1):
                draw.line(
                    [(x, header_y), (x, header_y + si['height'])],
                    fill=grid_color,
                )
                if ci < si['max_col']:
                    x += col_widths[ci]

            # 水平线
            for ri in range(len(rows_data) + 1):
                ry = header_y + ri * row_height
                draw.line(
                    [(10, ry), (10 + si['width'], ry)],
                    fill=grid_color,
                )

            y_offset += si['height'] + gap

            # 报告进度
            if progress_callback and total_sheets > 0:
                progress_callback(min(int((si_idx + 1) / total_sheets * 100), 99))

        # ---------- 保存 ----------
        output_ext = Path(output_path).suffix.lower()
        if output_ext in ('.jpg', '.jpeg'):
            img = img.convert('RGB')
            img.save(temp_path, 'JPEG', quality=95)
        else:
            img.save(temp_path, 'PNG')
    finally:
        img.close()

    finalize_file(temp_path, output_path)
    return output_path
